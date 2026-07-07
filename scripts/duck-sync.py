#!/usr/bin/env python3
"""
速卖通烤鸭贴纸客户 → 一键同步「飞书文档」+「网站地图/统计」。

用法:
    python3 scripts/duck-sync.py <速卖通导出.xlsx>            # 预览(dry-run)，不改任何东西
    python3 scripts/duck-sync.py <速卖通导出.xlsx> --apply    # 真正写飞书 + 网站 + git push
    python3 scripts/duck-sync.py <速卖通导出.xlsx> --apply --no-push   # 写飞书+网站，但不 git push
    python3 scripts/duck-sync.py <xlsx> --doc-md <飞书导出.md>  # 飞书 API 读不到时，用本地导出的 md 做去重基准

流程:
    1. 解析 xlsx → 过滤(只烤鸭贴纸 + 只已付款)
    2. 读取飞书现有客户(全文 fetch + 反转义) → 邮箱去重 → 得出真正新增
    3. 新城市 Nominatim 地理编码 → 追加 src/data/sticker-customers.json
    4. 飞书: 去重后全量重建整篇文档(v2 overwrite)——新客户插到本国表最前标🆕，
            重写顶部统计行 + 各国表 + 「📊 汇总」「🆕 本次新增」
    5. git commit & push (触发 Vercel)

依赖: openpyxl, lark-cli(已登录 --as user), curl
"""
import sys, os, re, json, subprocess, time, datetime

# ─── 配置 ────────────────────────────────────────────────────────────────
DUCK_ID = '1005010774387015'                  # 烤鸭贴纸商品 ID(只统计它)
DOC_TOKEN = 'FO1CdOry7owQ6Fx9qLzcWw88nCd'     # 飞书文档真实 docx token
EXCLUDE_STATUS = {'订单关闭', '等待买家付款', '交易关闭', '已取消', '等待付款'}
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SITE_JSON = os.path.join(ROOT, 'src', 'data', 'sticker-customers.json')

# 国家(收货国家原文/中文) → (中文名, 国旗, 网站用英文名)
COUNTRY = {
    'Japan': ('日本', '🇯🇵', 'Japan'), '日本': ('日本', '🇯🇵', 'Japan'),
    'Germany': ('德国', '🇩🇪', 'Germany'), '德国': ('德国', '🇩🇪', 'Germany'),
    'France': ('法国', '🇫🇷', 'France'), '法国': ('法国', '🇫🇷', 'France'),
    'South Korea': ('韩国', '🇰🇷', 'South Korea'), 'Korea, Republic of': ('韩国', '🇰🇷', 'South Korea'), '韩国': ('韩国', '🇰🇷', 'South Korea'),
    'Australia': ('澳大利亚', '🇦🇺', 'Australia'), '澳大利亚': ('澳大利亚', '🇦🇺', 'Australia'),
    'Switzerland': ('瑞士', '🇨🇭', 'Switzerland'), '瑞士': ('瑞士', '🇨🇭', 'Switzerland'),
    'Belgium': ('比利时', '🇧🇪', 'Belgium'), 'Netherlands': ('荷兰', '🇳🇱', 'Netherlands'),
    'Canada': ('加拿大', '🇨🇦', 'Canada'), '加拿大': ('加拿大', '🇨🇦', 'Canada'),
    'Spain': ('西班牙', '🇪🇸', 'Spain'), 'Italy': ('意大利', '🇮🇹', 'Italy'),
    'United Kingdom': ('英国', '🇬🇧', 'United Kingdom'), 'UK': ('英国', '🇬🇧', 'United Kingdom'),
    'Mexico': ('墨西哥', '🇲🇽', 'Mexico'), 'Malaysia': ('马来西亚', '🇲🇾', 'Malaysia'),
    'Portugal': ('葡萄牙', '🇵🇹', 'Portugal'), 'Poland': ('波兰', '🇵🇱', 'Poland'),
    'Singapore': ('新加坡', '🇸🇬', 'Singapore'), 'Finland': ('芬兰', '🇫🇮', 'Finland'),
    'Hungary': ('匈牙利', '🇭🇺', 'Hungary'), 'Chile': ('智利', '🇨🇱', 'Chile'),
    'Russia': ('俄罗斯', '🇷🇺', 'Russia'), 'Russian Federation': ('俄罗斯', '🇷🇺', 'Russia'),
    'United States': ('美国', '🇺🇸', 'United States'), 'USA': ('美国', '🇺🇸', 'United States'),
    'Austria': ('奥地利', '🇦🇹', 'Austria'), 'Sweden': ('瑞典', '🇸🇪', 'Sweden'),
    'Norway': ('挪威', '🇳🇴', 'Norway'), 'Denmark': ('丹麦', '🇩🇰', 'Denmark'),
    'Ireland': ('爱尔兰', '🇮🇪', 'Ireland'), 'New Zealand': ('新西兰', '🇳🇿', 'New Zealand'),
    'Czechia': ('捷克', '🇨🇿', 'Czechia'), 'Czech Republic': ('捷克', '🇨🇿', 'Czechia'),
    'Greece': ('希腊', '🇬🇷', 'Greece'), 'Brazil': ('巴西', '🇧🇷', 'Brazil'),
    'Uzbekistan': ('乌兹别克斯坦', '🇺🇿', 'Uzbekistan'), 'China': ('中国', '🇨🇳', 'China'), '中国': ('中国', '🇨🇳', 'China'),
    'Slovakia': ('斯洛伐克', '🇸🇰', 'Slovakia'), '斯洛伐克': ('斯洛伐克', '🇸🇰', 'Slovakia'),
}
# 飞书 parse 出的国家 key 是中文名——中文 key 缺失时全量重建会把该国旗子降级成 🏳️
for _v in list(COUNTRY.values()):
    COUNTRY.setdefault(_v[0], _v)
# 这些国家速卖通的"城市"列是行政区，真实城市在"扩展城市"列
EXT_CITY = {'Germany', '德国', 'Italy', '意大利', 'Poland', '波兰', 'Mexico', '墨西哥'}
STATUS_MAP = {'交易成功': '✅', '已完成': '✅', '等待买家收货': '等待收货',
              '卖家已发货': '已发货', '等待卖家发货': '待发货', '买家已付款': '已付款'}
HDR = ['序号', '收件人', '电话', '邮箱', '收货地址', '数量', '下单时间', '状态', '买家秀']


def log(msg): print(msg, flush=True)


def run(cmd, **kw):
    return subprocess.run(cmd, capture_output=True, text=True, cwd=ROOT, **kw)


def unescape(s):
    return re.sub(r'\\(.)', r'\1', s or '')


# ─── 1. 解析 xlsx ────────────────────────────────────────────────────────
def parse_xlsx(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))

    def c(r, i):
        v = r[i] if i < len(r) else None
        return str(v).strip() if v is not None else ''
    hdr = [c(rows[0], i) for i in range(len(rows[0]))]

    def idx(name):
        try:
            return hdr.index(name)
        except ValueError:
            return -1
    I = {k: idx(v) for k, v in {
        'status': '订单状态', 'time': '下单时间', 'pid': '商品 ID', 'prod': '商品信息',
        'fulladdr': '完整收货地址', 'recipient': '收件人名称', 'phone_cc': '联系电话',
        'mobile': '手机', 'email': '联系邮件', 'city': '城市',
        'extcity': '扩展城市（德/意/波/墨为真实的城市）', 'state': '州/省', 'country': '收货国家',
    }.items()}

    out = []
    for r in rows[1:]:
        if c(r, I['pid']) != DUCK_ID:
            continue
        if c(r, I['status']) in EXCLUDE_STATUS:
            continue
        cr = c(r, I['country'])
        cz, flag, en = COUNTRY.get(cr, (cr, '🏳️', cr))
        use_ext = cr in EXT_CITY or cz in ('德国', '意大利', '波兰', '墨西哥')
        city = c(r, I['extcity']) if (use_ext and c(r, I['extcity'])) else c(r, I['city'])
        if not city:
            city = c(r, I['extcity']) or c(r, I['state'])
        qty = 1
        m = re.search(r'商品数量:(\d+)', c(r, I['prod']))
        if m:
            qty = int(m.group(1))
        t = c(r, I['time'])
        dm = re.match(r'(\d{2})/(\d{2})/(\d{4})', t)
        md = f"{dm.group(1)}/{dm.group(2)}" if dm else t
        iso = f"{dm.group(3)}-{dm.group(1)}-{dm.group(2)}" if dm else ''
        addr = re.sub(r'[、，]', ', ', c(r, I['fulladdr'])).replace('|', '／')
        phone = (c(r, I['phone_cc']) + ' ' + c(r, I['mobile'])).strip()
        out.append({
            'email': c(r, I['email']).strip().lower(),
            'name': c(r, I['recipient']).replace('|', '／'),
            'phone': phone, 'address': addr, 'qty': qty,
            'date': md, 'iso': iso,
            'status': STATUS_MAP.get(c(r, I['status']), c(r, I['status'])),
            'country_zh': cz, 'flag': flag, 'country_en': en,
            'city': city, 'state': c(r, I['state']),
        })
    return out


# ─── 2. 读取飞书现有数据 ──────────────────────────────────────────────────
def fetch_doc_markdown(retries=4):
    """读飞书全文 markdown。网络不稳时重试；连续失败抛错(由调用方决定兜底)。"""
    last = ''
    for attempt in range(retries):
        p = run(['lark-cli', 'docs', '+fetch', '--doc', DOC_TOKEN, '--as', 'user',
                 '--offset', '0', '--limit', '99999', '--format', 'json'])
        raw = p.stdout
        try:
            j = json.loads(raw[raw.find('{'):])
        except Exception:
            last = raw[:200] or p.stderr[:200]
            time.sleep(1.5)
            continue
        md = j.get('data', j).get('markdown', '')
        if md and md.count('##') >= 15:
            return md
        last = j.get('error', {}).get('message', raw[:200])
        time.sleep(1.5)
    raise RuntimeError('读取飞书失败(已重试%d次): %s' % (retries, last))


def export_doc_markdown():
    """兜底: 用 drive +export 把文档「下载为 Markdown」到本地再读(等同飞书界面手动导出)。
    导出走的是异步任务通道，比 docs +fetch 更稳。返回 markdown 文本或抛错。"""
    outdir = os.path.join(ROOT, '.tmp_export')
    os.makedirs(outdir, exist_ok=True)
    last = ''
    for _ in range(3):
        p = run(['lark-cli', 'drive', '+export', '--token', DOC_TOKEN,
                 '--doc-type', 'docx', '--file-extension', 'markdown',
                 '--output-dir', '.tmp_export', '--overwrite', '--as', 'user'])
        try:
            j = json.loads(p.stdout[p.stdout.find('{'):])
            path = j.get('data', {}).get('saved_path')
            if path and os.path.exists(path):
                md = open(path, encoding='utf-8').read()
                if md.count('##') >= 15:
                    return md
        except Exception:
            last = (p.stdout or p.stderr)[:200]
        time.sleep(2)
    raise RuntimeError('export 兜底也失败: ' + last)


def parse_doc(md):
    """解析飞书文档 → {国家: [记录]}, 顺序保留。
    同时支持两种导出格式: 管道 markdown 表格(| a | b |) 和 lark-table HTML
    (飞书 API docs +fetch 返回的是后者，本地「下载为 Markdown」是前者)。
    """
    groups = {}
    for part in re.split(r'\n(?=##\s)', md):
        h = re.match(r'##\s*(.+)', part)
        if not h:
            continue
        title = h.group(1)
        if '汇总' in title or '新增' in title:
            continue
        cz = re.sub(r'[\U0001F1E6-\U0001F1FF]', '', title)
        cz = re.sub(r'（.*?）|\(.*?\)', '', cz).strip()
        groups.setdefault(cz, [])
        if '<lark-table' in part or '<table' in part:
            # 两种 HTML 表格: <lark-table><lark-tr><lark-td> (docs +fetch)
            # 和 <table><tr><td> (drive +export markdown)
            tr_tag = 'lark-tr' if '<lark-tr' in part else 'tr'
            td_tag = 'lark-td' if '<lark-td' in part else 'td'
            tbl_re = r'<lark-table[^>]*>(.*?)</lark-table>' if '<lark-table' in part else r'<table[^>]*>(.*?)</table>'
            for tbl in re.findall(tbl_re, part, re.S):
                rows = []
                for tr in re.findall(r'<%s[^>]*>(.*?)</%s>' % (tr_tag, tr_tag), tbl, re.S):
                    cells = [unescape(re.sub(r'<[^>]+>', '', re.sub(r'\s+', ' ', x)).strip())
                             for x in re.findall(r'<%s[^>]*>(.*?)</%s>' % (td_tag, td_tag), tr, re.S)]
                    rows.append(cells)
                if len(rows) < 2:
                    continue
                header = rows[0]
                for r in rows[1:]:
                    groups[cz].append(dict(zip(header, r)))
        else:
            header = None
            for ln in part.split('\n'):
                if not ln.strip().startswith('|'):
                    continue
                cells = [unescape(x.strip()) for x in ln.strip().strip('|').split('|')]
                if set(''.join(cells).replace(' ', '')) <= set('-:'):
                    continue
                if header is None:
                    header = cells
                    continue
                groups[cz].append(dict(zip(header, cells)))
    return groups


# ─── 3. 网站数据 ─────────────────────────────────────────────────────────
def geocode(city, country):
    q = f"{city}, {country}".replace(' ', '+')
    import urllib.parse
    url = 'https://nominatim.openstreetmap.org/search?q=' + urllib.parse.quote(f"{city}, {country}") + '&format=json&limit=1'
    p = run(['curl', '-s', url, '-H', 'User-Agent: TRACKID-Gallery/1.0'])
    try:
        d = json.loads(p.stdout)
        if d:
            return float(d[0]['lat']), float(d[0]['lon'])
    except Exception:
        pass
    return None


def jitter(seed):
    # 确定性抖动(基于邮箱hash)，避免重跑产生不同坐标
    h = sum(ord(c) for c in seed)
    return ((h % 200) - 100) / 10000.0, (((h * 7) % 200) - 100) / 10000.0


def update_website(new_custs, apply):
    data = json.load(open(SITE_JSON, encoding='utf-8'))
    before = len(data)
    existing_city = {d['city'].lower() for d in data}
    added = []
    for r in new_custs:
        city, en = r['city'], r['country_en']
        if not city or city.lower() in existing_city:
            continue
        coord = geocode(city, en)
        time.sleep(1.1)
        if not coord:
            log(f"   ⚠️  geocode 失败，跳过: {city}, {en}")
            continue
        dla, dln = jitter(r['email'] or city)
        rec = {'city': city, 'country': en,
               'lat': round(coord[0] + dla, 4), 'lng': round(coord[1] + dln, 4)}
        data.append(rec)
        existing_city.add(city.lower())
        added.append(rec)
        log(f"   + {city}, {en}  ({rec['lat']}, {rec['lng']})")
    if apply and added:
        json.dump(data, open(SITE_JSON, 'w', encoding='utf-8'), ensure_ascii=False, indent=2)
    cities = len({d['city'].lower() for d in data})
    countries = len({d['country'] for d in data})
    return {'before': before, 'after': len(data), 'added': len(added),
            'cities': cities, 'countries': countries}


# ─── 4. 飞书写入 ─────────────────────────────────────────────────────────
def esc(s):
    return (s or '').replace('|', '／').replace('\n', ' ').strip()


def render_table(cz, flag, recs):
    # 标题与表格之间不留空行：replace_range 不能跨段落，空行会触发
    # lark-cli「多段落」警告并偶发中止替换(2026-06 英/德两表实测踩坑)。
    out = [f'## {flag} {cz}（{len(recs)}人）',
           '|' + '|'.join(HDR) + '|', '|' + '|'.join(['---'] * len(HDR)) + '|']
    for i, r in enumerate(recs, 1):
        cells = [str(i)] + [esc(r.get(k, '')) for k in HDR[1:]]
        out.append('|' + '|'.join(cells) + '|')
    return '\n'.join(out) + '\n'


def detect_duplicate_sections(md):
    """返回文档中重复出现的章节标题（忽略汇总/新增章节）。"""
    counts = {}
    for t in re.findall(r'^##\s+.+', md, re.M):
        if '汇总' in t or '新增' in t:
            continue
        counts[t] = counts.get(t, 0) + 1
    return [t for t, n in counts.items() if n > 1]


def dedup_groups(raw_groups):
    """全局邮箱去重（保持首次出现顺序），去掉 🆕 标记。
    返回 (去重后 groups dict, unique_emails set)。"""
    seen = set()
    result = {}
    for cz, recs in raw_groups.items():
        clean = []
        for r in recs:
            email = unescape((r.get('邮箱', '') or '').strip().lower())
            if email and email in seen:
                continue
            if email:
                seen.add(email)
            r2 = dict(r)
            r2['收件人'] = r2.get('收件人', '').replace(' 🆕', '').strip()
            clean.append(r2)
        if clean:
            result[cz] = clean
    return result, seen


def overwrite_full_doc(groups, new_custs, apply):
    """把完整 groups（已含新客户行）以 overwrite 模式写回飞书。
    返回 (ok, err, total, items_sum, n_countries)。"""
    today = datetime.date.today().strftime('%Y-%m-%d')
    total = sum(len(v) for v in groups.values())
    items_sum = sum(int(re.sub(r'\D', '', r.get('数量', '1')) or 1)
                    for v in groups.values() for r in v)
    n_countries = len([k for k, v in groups.items() if v])

    parts = []
    stat = (f"Columbus Parody Bike Sticker Peking Duck（商品ID: {DUCK_ID}）"
            f"统计截至 {today} | 共 **{total}** 位客户 · **{items_sum}** 件 | **{n_countries}** 个国家")
    parts.append(stat)
    parts.append('')

    slovakia_key = None
    for cz, recs in groups.items():
        if not recs:
            continue
        if 'Slovakia' in cz or 'Slovak' in cz:
            slovakia_key = cz
            continue
        _, flag, _ = COUNTRY.get(cz, (cz, '🏳️', cz))
        parts.append(render_table(cz, flag, recs))
        parts.append('')

    dist = ' / '.join(
        f"{COUNTRY.get(cz,(cz,'🏳️',cz))[1]}{cz}{len(recs)}"
        for cz, recs in groups.items() if recs and 'Slovakia' not in cz and 'Slovak' not in cz
    )
    summary_tail = ''
    if new_custs:
        lines = '\n\n'.join(f"- {r['flag']} {r['name']}（{r['date']}, {r['qty']}件）"
                             for r in new_custs)
        summary_tail = f"\n### 🆕 本次新增 {len(new_custs)} 位客户：\n\n{lines}\n"
    summary_md = (
        f"## 📊 汇总\n\n"
        f"- **总客户数：{total} 人**（上次 {total - len(new_custs)} → 新增 {len(new_custs)}）\n\n"
        f"- **总销售件数：{items_sum} 件**\n\n"
        f"- **国家：{n_countries} 个**\n\n"
        f"- **国家分布：** {dist}\n"
        + summary_tail
    )
    parts.append(summary_md)
    parts.append('')

    if slovakia_key and groups.get(slovakia_key):
        _, flag, _ = COUNTRY.get(slovakia_key, (slovakia_key, '🏳️', slovakia_key))
        parts.append(render_table(slovakia_key, flag, groups[slovakia_key]))

    full_md = '\n'.join(parts)

    if not apply:
        return True, '(dry-run)', total, items_sum, n_countries

    tmp = os.path.join(ROOT, '.tmp_feishu', 'full_rebuild.md')
    os.makedirs(os.path.join(ROOT, '.tmp_feishu'), exist_ok=True)
    open(tmp, 'w', encoding='utf-8').write(full_md)

    # lark-cli ≥1.0.65 v2 接口: --command overwrite --doc-format markdown --content @file
    # (旧 --mode/--markdown v1 接口 2026-07 已下线)。@file 必须是相对 ROOT 的路径。
    p = run(['lark-cli', 'docs', '+update', '--doc', DOC_TOKEN, '--as', 'user',
             '--command', 'overwrite', '--doc-format', 'markdown',
             '--content', '@.tmp_feishu/full_rebuild.md'])
    ok = '"ok": true' in p.stdout and '"failed"' not in p.stdout
    err = ''
    if not ok:
        m = re.search(r'"message":\s*"([^"]+)"', p.stdout)
        err = m.group(1) if m else (p.stdout[:300] or p.stderr[:300])
    return ok, err, total, items_sum, n_countries


# ─── 主流程 ──────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    if not args or args[0] in ('-h', '--help'):
        print(__doc__)
        sys.exit(0)
    xlsx = args[0]
    apply = '--apply' in args
    no_push = '--no-push' in args
    if not os.path.exists(xlsx):
        log(f"❌ 找不到文件: {xlsx}")
        sys.exit(1)

    mode = '✍️  APPLY(真实写入)' if apply else '👀 DRY-RUN(仅预览, 加 --apply 才写入)'
    log(f"\n=== 烤鸭贴纸客户同步 [{mode}] ===\n")

    # 1. 解析 Excel
    parsed = parse_xlsx(xlsx)
    log(f"① Excel 解析: 已付款烤鸭订单 {len(parsed)} 条")

    # 2. 读飞书去重
    doc_md_path = None
    if '--doc-md' in args:
        doc_md_path = args[args.index('--doc-md') + 1]
    if doc_md_path:
        log(f"② 读取飞书基准(本地 md: {doc_md_path})…")
        md = open(doc_md_path, encoding='utf-8').read()
    else:
        log("② 读取飞书现有客户(API)…")
        try:
            md = fetch_doc_markdown()
        except RuntimeError:
            log("   docs +fetch 读不全，自动改用 drive 导出 Markdown 兜底…")
            try:
                md = export_doc_markdown()
                log("   ✓ 导出成功")
            except RuntimeError as e:
                log(f"\n❌ 飞书两种读取方式都失败: {e}")
                log("   为避免重复写入已中止。稍后重试，或手动在飞书「下载为 Markdown」后用 --doc-md <路径> 重跑。\n")
                sys.exit(2)
    doc_groups = {k: v for k, v in parse_doc(md).items()}

    # 安全闸: 读到的飞书客户数异常少 → 必是读取不全，停止(否则会把已存在客户当新增重复写入)
    _cur = sum(len(v) for v in doc_groups.values())
    if _cur < 50:
        log(f"\n❌ 只读到 {_cur} 位飞书客户，明显不全(正常应 90+)。为避免重复写入已中止。")
        log("   解决: 飞书「下载为 Markdown」后用 --doc-md <文件路径> 重跑。\n")
        sys.exit(2)

    # 重复章节检测：文档被意外复制两份时，同一章节标题出现多次
    dup_sections = detect_duplicate_sections(md)
    will_overwrite = bool(dup_sections)
    if dup_sections:
        log(f"   ⚠️  检测到 {len(dup_sections)} 个重复章节（如 {dup_sections[0]!r}），自动去重修复…")
        doc_groups, doc_emails = dedup_groups(doc_groups)
        cur_total = sum(len(v) for v in doc_groups.values())
        cur_items = sum(int(re.sub(r'\D', '', r.get('数量', '1')) or 1)
                        for v in doc_groups.values() for r in v)
        cur_countries = len([k for k, v in doc_groups.items() if v])
        log(f"   去重后: {cur_total} 人 / {cur_items} 件 / {cur_countries} 国 / {len(doc_emails)} 唯一邮箱")
    else:
        doc_emails = {(r.get('邮箱', '') or '').strip().lower()
                      for recs in doc_groups.values() for r in recs if r.get('邮箱', '').strip()}
        cur_total = sum(len(v) for v in doc_groups.values())
        cur_items = sum(int(re.sub(r'\D', '', r.get('数量', '1')) or 1)
                        for recs in doc_groups.values() for r in recs)
        cur_countries = len([k for k, v in doc_groups.items() if v])
        log(f"   飞书现有: {cur_total} 人 / {cur_items} 件 / {cur_countries} 国 / {len(doc_emails)} 唯一邮箱")

    # 去重(批内也去重)
    new_custs, dup, seen = [], [], set()
    for r in parsed:
        e = r['email']
        if e in doc_emails or e in seen:
            dup.append(r)
        else:
            seen.add(e)
            new_custs.append(r)
    log(f"③ 去重: 已在飞书 {len(dup)} 条 | ★ 真正新增 {len(new_custs)} 条")
    for r in new_custs:
        log(f"   ★ {r['name'][:22]:<24} {r['city'][:14]:<16} {r['country_zh']:<8} {r['email']}")
    if not new_custs:
        if will_overwrite:
            # 虽无新增客户，仍需修复文档重复问题
            log("\n✅ 没有新增客户，但检测到文档重复——执行去重修复。")
            log(f"\n⑤ 飞书文档（去重修复）…")
            ok, err, new_total, new_items, new_countries = overwrite_full_doc(doc_groups, [], apply)
            log(f"   {'✓ 修复成功' if ok else '✗ 修复失败'}: {new_total}人/{new_items}件/{new_countries}国"
                + (f"  → {err}" if not ok else ''))
            log(f"\n{'✅ 完成' if apply else '👀 预览结束 — 加 --apply 真正执行'}")
            return
        log("\n✅ 没有新增客户，无需更新。\n")
        return

    # 3. 网站
    log(f"\n④ 网站地图({'写入' if apply else '预览'})…")
    web = update_website(new_custs, apply)
    log(f"   脉冲点: {web['before']} → {web['after']} (+{web['added']}) | 城市 {web['cities']} | 国家 {web['countries']}")

    # 4. 飞书 — 一律去重后全量重建(overwrite)：
    #    lark-cli v2 已下线 replace_range，且历史上按章节 replace 两次把文档结构
    #    搞坏(章节复制/错并)，全量重建反而是两次验证过的稳妥路径。
    log(f"\n⑤ 飞书文档({'写入' if apply else '预览'})…")
    if will_overwrite:
        log(f"   (文档含重复章节，重建时一并去重)")
    merged = dict(doc_groups)
    for r in new_custs:
        cz = r['country_zh']
        merged.setdefault(cz, [])
        newrow = {'收件人': r['name'] + ' 🆕', '电话': r['phone'], '邮箱': r['email'],
                  '收货地址': r['address'], '数量': str(r['qty']),
                  '下单时间': r['date'], '状态': r['status'], '买家秀': ''}
        merged[cz].insert(0, newrow)
    ok, err, new_total, new_items, new_countries = overwrite_full_doc(merged, new_custs, apply)
    fb = [('全量重建', 'overwrite', ok, err)]

    for cz, op, ok, err in fb:
        log(f"   {'✓' if ok else '✗'} {cz} [{op}]" + ('' if ok else f"  → {err}"))
    log(f"   统计: 共{cur_total}→{new_total}人 · {cur_items}→{new_items}件 · {cur_countries}→{new_countries}国")

    # 5. git push
    if apply and web['added'] > 0:
        log(f"\n⑥ git commit & push…")
        if no_push:
            log("   --no-push: 跳过推送")
        else:
            run(['git', 'add', 'src/data/sticker-customers.json'])
            cities_str = ', '.join(sorted({r['city'] for r in new_custs}))
            run(['git', 'commit', '-q', '-m',
                 f"feat(stickers): add {web['added']} new buyer cities ({cities_str})\n\n"
                 f"Co-Authored-By: Claude Opus 4.8 (1M context) <noreply@anthropic.com>"])
            p = run(['git', 'push', 'origin', 'main'])
            log("   ✓ 已推送，Vercel 自动部署中" if p.returncode == 0 else f"   ✗ push 失败: {p.stderr[:200]}")

    log(f"\n{'✅ 同步完成' if apply else '👀 预览结束 — 确认无误后加 --apply 真正执行'}")
    log(f"   飞书: {cur_total}→{new_total}人 / {cur_items}→{new_items}件 / {new_countries}国")
    log(f"   网站: {web['before']}→{web['after']}个脉冲点\n")


if __name__ == '__main__':
    main()
